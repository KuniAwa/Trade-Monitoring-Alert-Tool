# -*- coding: utf-8 -*-
"""
フォルダ内のファイル名・階層を抽出し、簿記分類を推定してExcelに出力するスクリプト。
分類はExcel上で事後的に変更可能（ドロップダウンで選択）。
"""
import argparse
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any

# 簿記の6分類（事後変更用にExcelでドロップダウン表示）
CLASSIFICATIONS = ["①資産", "②負債", "③資本", "④収益", "⑤費用", "⑥その他"]

# ファイル名・フォルダ名から分類を推定するキーワード（必要に応じて追加可能）
KEYWORDS_BY_CATEGORY = {
    "①資産": [
        "現金", "預金", "売掛", "棚卸", "固定資産", "建物", "土地", "車両", "備品",
        "未収", "前払", "前払金", "貸付", "有価証券", "貯蔵品", "手形", "受取手形",
        "資産", "在庫", "繰越", "未収入金", "前払費用",
    ],
    "②負債": [
        "借入", "買掛", "未払", "預り", "前受", "支払手形", "社債", "負債",
        "未払金", "前受金", "預り金", "借入金", "未払費用", "前受収益",
    ],
    "③資本": [
        "資本金", "繰越", "剰余金", "資本準備", "利益準備", "資本", "純資産",
        "繰越利益", "自己資本",
    ],
    "④収益": [
        "売上", "売上高", "収益", "受取", "雑収入", "受取利息", "受取配当",
        "収入", "売上原価", "収益認識", "収益計上",
    ],
    "⑤費用": [
        "仕入", "人件費", "給与", "旅費", "消耗品", "支払", "経費", "外注",
        "租税", "減価償却", "広告", "交際費", "通信費", "光熱費", "諸会費",
        "費用", "原価", "仕入高", "販管費", "販売費", "一般管理費", "調査",
        "点検", "自主点検", "委託", "検収",
    ],
}


def infer_classification(name: str, parent_path: str) -> str:
    """
    ファイル名および親フォルダ名から簿記の6分類のいずれかを推定する。
    複数マッチした場合は最初にマッチした分類を返す（資産→負債→資本→収益→費用の順で優先）。
    """
    # 拡張子を除いた名前＋親パスを結合して検索（小文字化して比較）
    name_without_ext = Path(name).stem if "." in name else name
    # 親フォルダ名＋ファイル名（拡張子除く）の両方からキーワードを検索
    for cat in ["①資産", "②負債", "③資本", "④収益", "⑤費用"]:
        for kw in KEYWORDS_BY_CATEGORY.get(cat, []):
            if kw in name_without_ext or kw in parent_path:
                return cat
    return "⑥その他"


# サイズを読みやすい形式に変換（例: 1024 → "1.0 KB"）
def format_size(size_bytes: int) -> str:
    if size_bytes is None or size_bytes == 0:
        return ""
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if size_bytes < 1024:
            return f"{size_bytes:.1f} {unit}" if unit != "B" else f"{size_bytes} B"
        size_bytes /= 1024
    return f"{size_bytes:.1f} PB"

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def collect_folder_structure(root_path: Path) -> List[Dict[str, Any]]:
    """
    指定フォルダ以下を走査し、ファイル・フォルダの一覧を階層付きで収集する。
    """
    root = Path(root_path).resolve()
    if not root.is_dir():
        raise NotADirectoryError(f"指定パスはフォルダではありません: {root}")

    results = []

    for path in sorted(root.rglob("*")):
        try:
            rel = path.relative_to(root)
            depth = len(rel.parts)
            parent = str(rel.parent) if rel.parent != Path(".") else "(ルート)"
            name = path.name
            full_path = str(path)
            is_dir = path.is_dir()
            kind = "フォルダ" if is_dir else "ファイル"
            ext = path.suffix if path.is_file() and path.suffix else ""

            # サイズ・日時（stat で取得、権限エラー時は空）
            size_bytes = ""
            size_display = ""
            mtime_str = ""
            ctime_str = ""
            try:
                st = path.stat()
                if path.is_file():
                    size_bytes = st.st_size
                    size_display = format_size(st.st_size)
                else:
                    size_bytes = ""
                    size_display = "—"
                mtime_str = datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                ctime_str = datetime.fromtimestamp(st.st_ctime).strftime("%Y-%m-%d %H:%M:%S")
            except (PermissionError, OSError):
                pass

            # ファイル名・親フォルダから簿記分類を推定（事後的にExcelで変更可能）
            classification = infer_classification(name, parent)

            results.append({
                "階層レベル": depth,
                "親フォルダ": parent,
                "名前": name,
                "種類": kind,
                "分類": classification,
                "拡張子": ext,
                "サイズ(バイト)": size_bytes if size_bytes != "" else "",
                "サイズ(表示)": size_display,
                "作成日時": ctime_str,
                "更新日時": mtime_str,
                "フルパス": full_path,
            })
        except (PermissionError, OSError):
            continue

    # 階層・親・名前でソート
    results.sort(key=lambda x: (x["階層レベル"], x["親フォルダ"], x["名前"]))
    return results


def write_excel(data: List[Dict[str, Any]], output_path: Path) -> None:
    """収集したデータをExcelに書き出す。"""
    if not HAS_OPENPYXL:
        raise RuntimeError(
            "Excel出力には openpyxl が必要です。\n"
            "実行: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "フォルダ一覧"

    headers = ["階層レベル", "親フォルダ", "名前", "種類", "分類", "拡張子", "サイズ(バイト)", "サイズ(表示)", "作成日時", "更新日時", "フルパス"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="thin"),
            right=Side(style="thin") if col < len(headers) else Side(style="thin"),
        )

    for row, row_data in enumerate(data, 2):
        for col, key in enumerate(headers, 1):
            val = row_data.get(key, "")
            ws.cell(row=row, column=col, value=val)
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)

    # 「分類」列にドロップダウンを設定（事後的に変更可能）
    from openpyxl.worksheet.datavalidation import DataValidation
    col_classification = headers.index("分類") + 1  # 1-based
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(CLASSIFICATIONS)}"',
        allow_blank=False,
    )
    dv.error = "分類は一覧から選択してください"
    dv.errorTitle = "無効な入力"
    dv.prompt = "①資産～⑥その他のいずれかを選択"
    dv.promptTitle = "分類"
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col_classification)}2:{get_column_letter(col_classification)}{len(data) + 1}")

    # 列幅を調整
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["K"].width = 55

    wb.save(output_path)


def write_csv(data: List[Dict[str, Any]], output_path: Path) -> None:
    """openpyxlが無い場合の代替: CSVに出力（Excelで開ける）。"""
    import csv
    headers = ["階層レベル", "親フォルダ", "名前", "種類", "分類", "拡張子", "サイズ(バイト)", "サイズ(表示)", "作成日時", "更新日時", "フルパス"]
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        w.writeheader()
        w.writerows(data)


def main():
    parser = argparse.ArgumentParser(description="フォルダ階層をExcelに出力")
    parser.add_argument(
        "folder",
        nargs="?",
        default=".",
        help="対象フォルダのパス（省略時はカレントディレクトリ）",
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help="出力Excelファイル名（省略時は folder_structure_日時.xlsx）",
    )
    parser.add_argument(
        "--csv",
        action="store_true",
        help="Excelの代わりにCSVで出力する",
    )
    args = parser.parse_args()

    root = Path(args.folder).resolve()
    if not root.is_dir():
        print(f"エラー: フォルダが見つかりません: {root}")
        return 1

    print(f"対象フォルダ: {root}")
    print("収集中...")
    data = collect_folder_structure(root)
    print(f"取得件数: {len(data)} 件")

    if args.output:
        out_path = Path(args.output).resolve()
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = root / f"folder_structure_{timestamp}.xlsx"

    if args.csv:
        out_path = out_path.with_suffix(".csv")
        write_csv(data, out_path)
        print(f"CSVを保存しました: {out_path}")
    else:
        try:
            write_excel(data, out_path)
            print(f"Excelを保存しました: {out_path}")
        except RuntimeError as e:
            print(e)
            out_csv = out_path.with_suffix(".csv")
            write_csv(data, out_csv)
            print(f"代わりにCSVを保存しました: {out_csv}")

    return 0


if __name__ == "__main__":
    exit(main())
